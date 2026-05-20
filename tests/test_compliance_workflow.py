from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlmodel import Session, select

from gmail_automation.dao import ImportDAO, TemplateDAO, WorkflowDAO
from gmail_automation.db import init_db
from gmail_automation.db_models import Client, ClientQuarter, Counterparty, CounterpartyField, DocumentJob, EmailMessage, GeneratedDocument, Template
from gmail_automation.docx_utils import build_docx_from_paragraphs, extract_docx_text
from gmail_automation.liquid_utils import extract_liquid_variables, render_liquid_template
from gmail_automation.modern_ui import ModernComplianceController, _attach_file_picker, _load_smtp_password, build_modern_ui_controls
from gmail_automation.services import ClientService, DashboardService, ImportService, TemplateService, WorkflowService
from gmail_automation.settings_store import SettingsService


def test_liquid_variable_extraction_and_strict_rendering() -> None:
    text = "Dear {{ party_name }}, balance {{ row.balance }} for {{ quarter.name }}"

    assert extract_liquid_variables(text) == {"party_name", "row.balance", "quarter.name"}
    rendered = render_liquid_template(
        text,
        {
            "party_name": "ABC Lender",
            "row": {"balance": "1000"},
            "quarter": {"name": "Q1"},
        },
    )

    assert rendered.text == "Dear ABC Lender, balance 1000 for Q1"


def test_excel_column_style_placeholder_rendering() -> None:
    text = "Dear {{Party Name}}, balance {{ Balance }} for {{ row.party_name }}"

    assert extract_liquid_variables(text) == {"Party Name", "Balance", "row.party_name"}
    rendered = render_liquid_template(
        text,
        {
            "Party Name": "ABC Lender",
            "Balance": "1000",
            "row": {"party_name": "ABC Lender"},
        },
    )

    assert rendered.text == "Dear ABC Lender, balance 1000 for ABC Lender"


def test_excel_import_preserves_standard_and_dynamic_fields(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _sample_workbook(tmp_path)

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        summary = ImportService(session).import_excel(client.id, quarter.id, workbook_path)
        rows = ImportDAO(session).list_counterparties(quarter.id)
        fields = session.exec(select(CounterpartyField).where(CounterpartyField.counterparty_id == rows[0].id)).all()

    assert summary.rows_imported == 2
    assert summary.sheet_counts == {"Creditors": 2}
    assert rows[0].party_name == "Alpha Finance"
    assert {field.field_name for field in fields} >= {"Party Name", "Email To(Address)", "Balance", "Custom Ref"}


def test_excel_import_status_columns_seed_compliance_summary(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _status_workbook(tmp_path)

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        ImportService(session).import_excel(client.id, quarter.id, workbook_path)
        rows = session.exec(select(Counterparty).where(Counterparty.client_quarter_id == quarter.id)).all()
        fields = {
            row.party_name: ImportDAO(session).fields_for_counterparty(row.id)
            for row in rows
            if row.id is not None
        }
        summary = DashboardService(session).compliance_summary(quarter.id)

    assert {row.party_name: row.status for row in rows} == {
        "Alpha Finance": "non_compliant",
        "Beta Bank": "non_compliant",
        "Gamma Capital": "non_compliant",
    }
    assert fields["Alpha Finance"]["Status"] == "Sent"
    assert fields["Beta Bank"]["ReadyToSend"] == "Y"
    assert fields["Gamma Capital"]["BounceReceived"] == "Yes"
    assert summary.fully_compliant == 0
    assert summary.partially_compliant == 0
    assert summary.non_compliant == 3
    assert summary.party_type_counts == {"Creditor": 3}


def test_excel_import_skips_readme_or_non_counterparty_sheets(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _workbook_with_readme_sheet(tmp_path)

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        summary = ImportService(session).import_excel(client.id, quarter.id, workbook_path)
        rows = ImportDAO(session).list_counterparties(quarter.id)

    assert summary.rows_imported == 2
    assert summary.sheet_counts == {"Tracking": 2}
    assert [row.party_name for row in rows] == ["Alpha Finance", "Beta Bank"]
    assert all(row.source_sheet == "Tracking" for row in rows)


def test_templates_mappings_generation_and_dashboard_queries(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _sample_workbook(tmp_path)

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        ImportService(session).import_excel(client.id, quarter.id, workbook_path)

        template_service = TemplateService(session)
        template_service.save_text_template(quarter.id, "document", "Confirmation Text", "Hello {{ party_name }}: {{ balance }}")
        template_service.save_text_template(quarter.id, "mail_subject", "Subject", "Confirm {{ party_name }}")
        template_service.save_text_template(quarter.id, "mail_body", "Body", "Balance is {{ balance }}")
        template_service.save_mappings(
            quarter.id,
            {
                "party_name": ("excel_column", "Party Name", ""),
                "balance": ("excel_column", "Balance", ""),
            },
        )

        assert template_service.validate_mappings(quarter.id, {"Party Name", "Balance"}) == []
        assert TemplateDAO(session).variables_for_quarter(quarter.id) == {"party_name", "balance"}

        workflow = WorkflowService(session, output_root=tmp_path / "generated")
        jobs = workflow.enqueue_document_generation(quarter.id)
        generated = workflow.generate_pending_documents(quarter.id)
        messages = workflow.queue_email_messages(quarter.id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        sent = workflow.mark_preview_batch_sent(quarter.id)
        summary = DashboardService(session).compliance_summary(quarter.id)
        job_counts = WorkflowDAO(session).generated_document_counts(quarter.id)
        email_counts = WorkflowDAO(session).email_status_counts(quarter.id)

        persisted_jobs = session.exec(select(DocumentJob).where(DocumentJob.client_quarter_id == quarter.id)).all()
        persisted_messages = session.exec(select(EmailMessage).where(EmailMessage.client_quarter_id == quarter.id)).all()
        persisted_counterparties = session.exec(select(Counterparty).where(Counterparty.client_quarter_id == quarter.id)).all()

    assert len(jobs) == 2
    assert len(generated) == 2
    assert all(result.status == "generated" for result in generated)
    assert len(messages) == 2
    assert sent == 2
    assert summary.fully_compliant == 0
    assert summary.non_compliant == 2
    assert job_counts["generated"] == 2
    assert email_counts["preview_sent"] == 2
    assert {job.status for job in persisted_jobs} == {"generated"}
    assert {message.status for message in persisted_messages} == {"preview_sent"}
    assert {counterparty.status for counterparty in persisted_counterparties} == {"non_compliant"}


def test_document_template_files_generate_for_all_counterparties(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    annexure_path = tmp_path / "annexure.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")
    annexure_path.write_text("Annexure for {{ row.party_name }}", encoding="utf-8")

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        ImportService(session).import_excel(client.id, quarter.id, workbook_path)

        template_service = TemplateService(session)
        templates = template_service.save_document_templates(
            quarter.id,
            [letter_path, annexure_path],
            storage_dir=tmp_path / "template_store",
            client_id=client.id,
        )
        template_service.save_mappings(
            quarter.id,
            {
                "party_name": ("excel_column", "Party Name", ""),
                "balance": ("excel_column", "Balance", ""),
                "row.party_name": ("counterparty_field", "party_name", ""),
            },
        )

        workflow = WorkflowService(session, output_root=tmp_path / "generated")
        jobs = workflow.enqueue_document_generation(quarter.id)
        generated = workflow.generate_pending_documents(quarter.id)
        documents = session.exec(select(GeneratedDocument)).all()
        variables = TemplateDAO(session).variables_for_quarter(quarter.id)

    assert len(templates) == 2
    assert variables == {"party_name", "balance", "row.party_name"}
    assert len(jobs) == 4
    assert len(generated) == 4
    assert all(result.status == "generated" for result in generated)
    assert len(documents) == 4
    assert any("Alpha Finance" in Path(result.file_path).read_text(encoding="utf-8") for result in generated)


def test_pdf_authorisation_attachment_variables_validate_against_excel_columns(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _sample_workbook(tmp_path)
    authorisation_pdf = tmp_path / "Authorisation for Direct Balance Confirmation.pdf"
    authorisation_pdf.write_bytes(b"%PDF-1.4\nAuthorization for {{Party Name}}\n%%EOF")

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        ImportService(session).import_excel(client.id, quarter.id, workbook_path)
        template_service = TemplateService(session)
        templates = template_service.save_document_templates(
            quarter.id,
            [authorisation_pdf],
            storage_dir=tmp_path / "template_store",
            client_id=client.id,
        )
        template_names = [template.name for template in templates]
        assert TemplateDAO(session).variables_for_quarter(quarter.id) == {"Party Name"}
        template_service.save_mappings(quarter.id, {"Party Name": ("excel_column", "Party Name", "")})
        mapping_errors = template_service.validate_mappings(quarter.id, {"Party Name"})
        workflow = WorkflowService(session, output_root=tmp_path / "generated")
        jobs = workflow.enqueue_document_generation(quarter.id)
        generated = workflow.generate_pending_documents(quarter.id)

    assert template_names == ["Authorisation for Direct Balance Confirmation"]
    assert mapping_errors == []
    assert len(jobs) == 2
    assert len(generated) == 2
    assert all(Path(result.file_path).read_bytes() == authorisation_pdf.read_bytes() for result in generated)


def test_modern_ui_controller_and_controls_smoke(tmp_path: Path) -> None:
    """Flet smoke test without a Page. Manual: run launch_modern_ui(), Configure → Browse… or Import Excel (opens native picker on desktop); status shows success or Error: …"""
    ft = pytest.importorskip("flet")
    controller = ModernComplianceController(tmp_path / "modern.db")
    try:
        created = controller.create_client_record("Purple United")
        ui = build_modern_ui_controls(ft, controller)

        ui["refresh"]()
        ui["show_page"]("Clients")
        cards = controller.client_cards()

        assert len(ui["controls"]) >= 4
        assert ui["selected_quarter"].value is None
        assert cards[0]["client_name"] == "Purple United"
        assert cards[0]["quarter_label"] == "No quarter"
        with pytest.raises(ValueError, match="Create a quarter"):
            controller.current_quarter_id_for_client(created["id"])

        controller.create_quarter(created["id"], "2026-27", "Q1")
        ui["refresh"]()
        assert controller.current_quarter_id_for_client(created["id"]) == int(ui["selected_quarter"].value)
    finally:
        controller.close()


def test_modern_controller_import_workbook_validates_path_and_client(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    try:
        a = controller.create_client_record("Client A")
        controller.create_quarter(a["id"], "2026-27", "Q1")
        quarter_a = controller.current_quarter_id_for_client(a["id"])
        b = controller.create_client_record("Client B")
        controller.create_quarter(b["id"], "2026-27", "Q1")
        quarter_b = controller.current_quarter_id_for_client(b["id"])

        with pytest.raises(ValueError, match="Workbook not found"):
            controller.import_workbook(quarter_a, str(tmp_path / "missing.xlsx"), client_id=a["id"])
        with pytest.raises(ValueError, match="does not belong"):
            controller.import_workbook(quarter_b, str(workbook_path), client_id=a["id"])
        assert controller.import_workbook(quarter_a, str(workbook_path), client_id=a["id"]).startswith("Imported")
    finally:
        controller.close()


def test_modern_controller_import_reports_legacy_status_wiring(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _status_workbook(tmp_path)
    try:
        created = controller.create_client_record("Client A")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])

        message = controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        summary = controller.get_client_compliance_summary(created["id"])["summary"]

        assert "Legacy status columns wired: BounceReceived, ReadyToSend, Status." in message
        assert summary.fully_compliant == 0
        assert summary.partially_compliant == 0
        assert summary.non_compliant == 3
        assert summary.party_type_counts == {"Creditor": 3}
    finally:
        controller.close()


def test_modern_controller_excel_persistence_overwrites_stale_quarter_rows_and_workflow(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    first_workbook = _sample_workbook(tmp_path)
    replacement_workbook = _single_row_workbook(tmp_path, "replacement.xlsx", "Gamma Capital", "gamma@example.com", "3000")
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(first_workbook), client_id=created["id"])
        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        controller.save_document_templates(quarter_id, str(letter_path))
        assert controller.auto_map_variables(quarter_id) == "Mappings valid."
        controller.run_generation(quarter_id)
        controller.queue_and_preview_send(quarter_id)

        assert len(controller.quarter_counterparty_statuses(quarter_id)) == 2

        message = controller.import_workbook(quarter_id, str(replacement_workbook), client_id=created["id"])
        rows = controller.quarter_counterparty_statuses(quarter_id)

        assert message == "Imported 1 row(s) from 1 sheet(s)."
        assert [row["party_name"] for row in rows] == ["Gamma Capital"]
        assert rows[0]["document_status"] == "not generated"
        assert rows[0]["mail_status"] == "not generated"
    finally:
        controller.close()


def test_modern_ui_file_picker_attaches_to_flet_services() -> None:
    ft = pytest.importorskip("flet")

    class DummyPage:
        def __init__(self) -> None:
            self.services = []
            self.overlay = []

    page = DummyPage()
    picker = _attach_file_picker(ft, page)

    assert picker in page.services
    assert page.overlay == []


def test_modern_controller_configure_flow_import_templates_mapping_and_summary(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        assert controller.client_cards()[0]["quarter_id"] is None

        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        import_message = controller.import_workbook(quarter_id, str(workbook_path))
        template_message = controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        document_message = controller.save_document_templates(quarter_id, str(letter_path))
        mapping_message = controller.auto_map_variables(quarter_id)
        imported_summary = controller.get_client_compliance_summary(created["id"])

        generation_message = controller.run_generation(quarter_id)
        queue_message = controller.queue_and_preview_send(quarter_id)
        completed_summary = controller.get_client_compliance_summary(created["id"])

        assert import_message == "Imported 2 row(s) from 1 sheet(s)."
        assert "party_name" in template_message
        assert "Saved 1 document template" in document_message
        assert mapping_message == "Mappings valid."
        assert imported_summary["summary"].total == 2
        assert imported_summary["summary"].non_compliant == 2
        assert generation_message == "Generated 2 document job(s); 0 need attention."
        assert queue_message == "Queued 2 email(s); marked 2 as preview (not delivered)."
        assert completed_summary["summary"].fully_compliant == 0
        assert completed_summary["summary"].email_counts["preview_sent"] == 2
        assert completed_summary["summary"].generation_counts["generated"] == 2
    finally:
        controller.close()


def test_modern_controller_row_level_compliance_detail_tracks_workflow_steps(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])

        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        imported_rows = controller.quarter_counterparty_statuses(quarter_id)

        assert [row["party_name"] for row in imported_rows] == ["Alpha Finance", "Beta Bank"]
        assert {row["document_status"] for row in imported_rows} == {"not generated"}
        assert {row["compliance_status"] for row in imported_rows} == {"non_compliant"}
        assert {row["mail_status"] for row in imported_rows} == {"not generated"}
        assert {row["mail_sent"] for row in imported_rows} == {"No"}

        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        controller.save_document_templates(quarter_id, str(letter_path))
        assert controller.auto_map_variables(quarter_id) == "Mappings valid."
        controller.run_generation(quarter_id)
        controller.queue_and_preview_send(quarter_id)

        completed_rows = controller.quarter_counterparty_statuses(quarter_id)

        assert {row["document_status"] for row in completed_rows} == {"generated (1)"}
        assert {row["document_count"] for row in completed_rows} == {1}
        assert {row["compliance_status"] for row in completed_rows} == {"non_compliant"}
        assert {row["mail_status"] for row in completed_rows} == {"preview_sent"}
        assert {row["mail_sent"] for row in completed_rows} == {"No"}
    finally:
        controller.close()


def test_modern_controller_selected_workflow_retrigger_affects_only_selected_rows(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        controller.save_document_templates(quarter_id, str(letter_path))
        assert controller.auto_map_variables(quarter_id) == "Mappings valid."

        rows = controller.quarter_counterparty_statuses(quarter_id)
        alpha_id = next(row["counterparty_id"] for row in rows if row["party_name"] == "Alpha Finance")

        with pytest.raises(ValueError, match="Generate documents"):
            controller.queue_and_preview_send_selected(quarter_id, {alpha_id})
        assert controller.regenerate_documents(quarter_id, {alpha_id}) == "Regenerated 1 selected document job(s); 0 need attention."
        with Session(controller.engine) as session:
            session.add(
                EmailMessage(
                    client_quarter_id=quarter_id,
                    counterparty_id=alpha_id,
                    to_email="alpha@example.com",
                    subject="Old failed message",
                    status="failed",
                    attempts=4,
                    retry_locked=True,
                    next_retry_at="2099-01-01 00:00:00",
                    error="old authentication failure",
                )
            )
            session.commit()
        assert controller.queue_and_preview_send_selected(quarter_id, {alpha_id}) == (
            "Queued 1 selected email(s); marked 1 as preview (not delivered)."
        )

        completed_rows = {row["party_name"]: row for row in controller.quarter_counterparty_statuses(quarter_id)}
        with Session(controller.engine) as session:
            alpha_messages = list(session.exec(select(EmailMessage).where(EmailMessage.counterparty_id == alpha_id)))

        assert completed_rows["Alpha Finance"]["document_status"] == "generated (1)"
        assert completed_rows["Alpha Finance"]["compliance_status"] == "non_compliant"
        assert completed_rows["Alpha Finance"]["mail_status"] == "preview_sent"
        assert completed_rows["Alpha Finance"]["mail_sent"] == "No"
        assert {message.retry_locked for message in alpha_messages} == {False}
        assert {message.error for message in alpha_messages} == {""}
        assert completed_rows["Beta Bank"]["document_status"] == "not generated"
        assert completed_rows["Beta Bank"]["compliance_status"] == "non_compliant"
        assert completed_rows["Beta Bank"]["mail_status"] == "not generated"
        assert completed_rows["Beta Bank"]["mail_sent"] == "No"
    finally:
        controller.close()


def test_modern_controller_doc_regeneration_readiness_not_blocked_by_mail_only_mapping_errors(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_document_templates(quarter_id, str(letter_path))
        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance for {{ quarter.name }}")
        controller.auto_map_variables(quarter_id)
        readiness = controller.quarter_workflow_readiness(quarter_id)
        row_ids = {row["counterparty_id"] for row in controller.quarter_counterparty_statuses(quarter_id)}

        assert readiness["can_generate_documents"] is True
        assert readiness["mail_mapping_errors"] == []
        assert readiness["can_send_mail"] is readiness["send_enabled"]
        assert controller.regenerate_documents(quarter_id, row_ids).startswith("Regenerated")
    finally:
        controller.close()


def test_modern_controller_reset_status_autofills_missing_document_mappings(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    doc_path = tmp_path / "doc.txt"
    doc_path.write_text("Hello {{ Party Name }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Org Two")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_document_templates(quarter_id, str(doc_path))
        controller.save_mail_templates(quarter_id, "Sub {{ party_name }}", "Body {{ balance }}")
        controller.auto_map_variables(quarter_id)

        # Simulate stale mapping state that misses the newly required "Party Name" variable.
        with Session(controller.engine) as session:
            TemplateService(session).save_mappings(quarter_id, {"party_name": ("excel_column", "Party Name", "")})

        before = controller.quarter_workflow_readiness(quarter_id)
        assert before["document_mapping_errors"] == []
        assert before["can_generate_documents"] is True

        message = controller.reset_quarter_after_config_change(quarter_id, reset_documents=True)
        after = controller.quarter_workflow_readiness(quarter_id)

        assert message.startswith("Reset 2 counterparty row(s):")
        assert after["document_mapping_errors"] == []
        assert after["can_generate_documents"] is True
    finally:
        controller.close()


def test_modern_controller_send_mail_selected_requires_send_mode_and_credentials(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_document_templates(quarter_id, str(letter_path))
        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        controller.auto_map_variables(quarter_id)
        row_ids = {row["counterparty_id"] for row in controller.quarter_counterparty_statuses(quarter_id)}
        controller.regenerate_documents(quarter_id, row_ids)

        with pytest.raises(ValueError, match="Enable send mode"):
            controller.send_mail_selected(quarter_id, row_ids)
    finally:
        controller.close()


def test_modern_controller_mail_builtin_variables_do_not_require_mapping(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_document_templates(quarter_id, str(letter_path))
        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance {{ balance }} for {{ quarter.name }}")

        readiness = controller.quarter_workflow_readiness(quarter_id)

        assert readiness["mail_mapping_errors"] == []
    finally:
        controller.close()


def test_modern_controller_mail_template_preview_uses_latest_active_template(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.save_mail_templates(quarter_id, "Original subject", "Original body")

        with Session(controller.engine) as session:
            dao = TemplateDAO(session)
            dao.create_template(
                template_type="mail_subject",
                name="Legacy Subject",
                version="legacy-sub",
                checksum="legacy-sub",
                client_quarter_id=quarter_id,
                content_text="Legacy subject should not win",
            )
            dao.create_template(
                template_type="mail_body",
                name="Legacy Body",
                version="legacy-body",
                checksum="legacy-body",
                client_quarter_id=quarter_id,
                content_text="Legacy body should not win",
            )

        controller.save_mail_templates(quarter_id, "Newest subject", "Newest body")
        preview = controller.preview_mail_templates(quarter_id)

        assert preview == {"subject": "Newest subject", "body": "Newest body"}
    finally:
        controller.close()


def test_modern_controller_mail_queue_uses_saved_templates_per_client_and_quarter(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}", encoding="utf-8")
    try:
        client_a = controller.create_client_record("Client A")
        controller.create_quarter(client_a["id"], "2026-27", "Q1")
        quarter_a = controller.current_quarter_id_for_client(client_a["id"])
        controller.import_workbook(quarter_a, str(workbook_path), client_id=client_a["id"])
        controller.save_document_templates(quarter_a, str(letter_path))
        controller.save_mail_templates(quarter_a, "A subject {{ party_name }}", "A body {{ balance }}")
        controller.auto_map_variables(quarter_a)
        controller.run_generation(quarter_a)

        client_b = controller.create_client_record("Client B")
        controller.create_quarter(client_b["id"], "2026-27", "Q1")
        quarter_b = controller.current_quarter_id_for_client(client_b["id"])
        controller.import_workbook(quarter_b, str(workbook_path), client_id=client_b["id"])
        controller.save_document_templates(quarter_b, str(letter_path))
        controller.save_mail_templates(quarter_b, "B subject {{ party_name }}", "B body {{ balance }}")
        controller.auto_map_variables(quarter_b)
        controller.run_generation(quarter_b)

        assert controller.queue_and_preview_send(quarter_a) == "Queued 2 email(s); marked 2 as preview (not delivered)."
        assert controller.queue_and_preview_send(quarter_b) == "Queued 2 email(s); marked 2 as preview (not delivered)."

        with Session(controller.engine) as session:
            messages_a = list(session.exec(select(EmailMessage).where(EmailMessage.client_quarter_id == quarter_a)))
            messages_b = list(session.exec(select(EmailMessage).where(EmailMessage.client_quarter_id == quarter_b)))

        assert len(messages_a) == 2
        assert len(messages_b) == 2
        assert all(message.subject.startswith("A subject ") for message in messages_a)
        assert all(message.body.startswith("A body ") for message in messages_a)
        assert all(message.subject.startswith("B subject ") for message in messages_b)
        assert all(message.body.startswith("B body ") for message in messages_b)
    finally:
        controller.close()


def test_modern_controller_document_reads_are_scoped_to_client_quarter(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}", encoding="utf-8")
    try:
        client_a = controller.create_client_record("Client A")
        controller.create_quarter(client_a["id"], "2026-27", "Q1")
        quarter_a = controller.current_quarter_id_for_client(client_a["id"])
        controller.import_workbook(quarter_a, str(workbook_path), client_id=client_a["id"])
        controller.save_document_templates(quarter_a, str(letter_path))
        controller.auto_map_variables(quarter_a)
        controller.run_generation(quarter_a)

        client_b = controller.create_client_record("Client B")
        controller.create_quarter(client_b["id"], "2026-27", "Q1")
        quarter_b = controller.current_quarter_id_for_client(client_b["id"])
        controller.import_workbook(quarter_b, str(workbook_path), client_id=client_b["id"])

        rows_b = controller.quarter_counterparty_statuses(quarter_b)
        target_b_id = rows_b[0]["counterparty_id"]

        with Session(controller.engine) as session:
            job_a = session.exec(select(DocumentJob).where(DocumentJob.client_quarter_id == quarter_a)).first()
            assert job_a is not None and job_a.id is not None
            session.add(
                GeneratedDocument(
                    document_job_id=job_a.id,
                    counterparty_id=target_b_id,
                    template_id=job_a.template_id,
                    file_type="txt",
                    file_path=str(letter_path),
                    checksum="corrupt-row",
                    template_version="test",
                )
            )
            session.commit()

        refreshed_rows_b = controller.quarter_counterparty_statuses(quarter_b)
        leaked_docs = controller.generated_documents_for_counterparty(quarter_b, target_b_id)

        assert all(row["document_count"] == 0 for row in refreshed_rows_b)
        assert all(row["document_status"] == "not generated" for row in refreshed_rows_b)
        assert leaked_docs == []
    finally:
        controller.close()


def test_modern_controller_new_client_without_quarter_is_listed_for_configuration(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    try:
        seeded = controller.create_client_record("Purple United")
        controller.create_quarter(seeded["id"], "2026-27", "Q1")
        controller.create_client_record("acme")

        cards = controller.client_cards()
        acme_card = next(card for card in cards if card["client_name"] == "acme")
        assert acme_card["quarter_id"] is None
        assert any(card["client_name"] == "Purple United" for card in cards)
        assert any(card["client_name"] == "acme" for card in cards)
    finally:
        controller.close()


def test_modern_controller_mail_settings_roundtrip(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    try:
        message = controller.save_mail_settings(
            mail_provider="gmail_smtp",
            send_mode="send",
            sender_email="sender@example.com",
            fallback_providers="webtel_smtp",
            daily_send_limit=500,
            per_email_delay_seconds=2,
            smtp_host="smtp.gmail.com",
            smtp_port=587,
            smtp_username="sender@example.com",
            smtp_password="secret",
        )
        settings = controller.get_mail_settings()

        assert message == "Mail settings saved."
        assert settings["mail_provider"] == "gmail_smtp"
        assert settings["send_mode"] == "send"
        assert settings["sender_email"] == "sender@example.com"
        assert settings["fallback_providers"] == "webtel_smtp"
        assert settings["daily_send_limit"] == "500"
        assert settings["smtp_password_saved"] is True
        with Session(controller.engine) as session:
            assert _load_smtp_password(SettingsService(session), "gmail_smtp", "sender@example.com") == "secret"
    finally:
        controller.close()


def test_modern_controller_add_and_delete_document_template_roundtrip(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    first = tmp_path / "first.txt"
    second = tmp_path / "second.txt"
    first.write_text("First {{ party_name }}", encoding="utf-8")
    second.write_text("Second {{ party_name }}", encoding="utf-8")
    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.save_document_templates(quarter_id, str(first))
        controller.add_document_templates(quarter_id, str(second))

        templates = controller.list_document_templates(quarter_id)
        names = {row["name"] for row in templates}
        assert names == {"first", "second"}

        first_template_id = next(row["id"] for row in templates if row["name"] == "first")
        assert controller.delete_document_template(first_template_id).startswith("Deleted document template:")

        remaining = controller.list_document_templates(quarter_id)
        assert len(remaining) == 1
        assert remaining[0]["name"] == "second"
    finally:
        controller.close()


def test_modern_controller_generated_document_preview_for_selected_counterparty(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")
    note_path = tmp_path / "note.txt"
    note_path.write_text("Note for {{ party_name }}: {{ balance }}", encoding="utf-8")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_document_templates(quarter_id, f"{letter_path}\n{note_path}")
        assert controller.auto_map_variables(quarter_id) == "Mappings valid."
        assert controller.regenerate_documents(quarter_id, {row["counterparty_id"] for row in controller.quarter_counterparty_statuses(quarter_id)}).startswith(
            "Regenerated"
        )

        rows = controller.quarter_counterparty_statuses(quarter_id)
        alpha_id = next(row["counterparty_id"] for row in rows if row["party_name"] == "Alpha Finance")
        generated_docs = controller.generated_documents_for_counterparty(quarter_id, alpha_id)
        preview = controller.preview_generated_document(generated_docs[0]["id"])

        assert len(generated_docs) == 2
        assert next(row for row in rows if row["party_name"] == "Alpha Finance")["document_count"] == 2
        assert len(next(row for row in rows if row["party_name"] == "Alpha Finance")["documents"]) == 2
        assert generated_docs[0]["file_path"].endswith(".txt")
        assert "Alpha Finance" in preview
        assert "1000" in preview
        assert "Alpha Finance" in controller.preview_generated_document(generated_docs[1]["id"])
    finally:
        controller.close()


def test_modern_controller_delete_client_removes_associated_workflow_and_config_data(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")

    try:
        victim = controller.create_client_record("Delete Me")
        controller.create_quarter(victim["id"], "2026-27", "Q1")
        victim_quarter_id = controller.current_quarter_id_for_client(victim["id"])
        controller.import_workbook(victim_quarter_id, str(workbook_path), client_id=victim["id"])
        controller.save_document_templates(victim_quarter_id, str(letter_path))
        controller.save_mail_templates(victim_quarter_id, "Confirm {{ party_name }}", "Balance is {{ balance }}")
        controller.auto_map_variables(victim_quarter_id)
        controller.run_generation(victim_quarter_id)
        controller.queue_and_preview_send(victim_quarter_id)
        generated_dir = tmp_path / "generated" / str(victim_quarter_id)
        template_dir = tmp_path / "templates" / str(victim_quarter_id)
        assert generated_dir.exists()
        assert template_dir.exists()

        survivor = controller.create_client_record("Keep Me")
        controller.create_quarter(survivor["id"], "2026-27", "Q1")
        survivor_quarter_id = controller.current_quarter_id_for_client(survivor["id"])
        controller.import_workbook(survivor_quarter_id, str(workbook_path), client_id=survivor["id"])

        message = controller.delete_client_record(victim["id"])

        with Session(controller.engine) as session:
            victim_client = session.get(Client, victim["id"])
            victim_rows = session.exec(select(Counterparty).where(Counterparty.client_id == victim["id"])).all()
            victim_templates = session.exec(
                select(Template).where((Template.client_id == victim["id"]) | (Template.client_quarter_id == victim_quarter_id))
            ).all()
            victim_quarters = session.exec(select(ClientQuarter).where(ClientQuarter.client_id == victim["id"])).all()
            victim_jobs = session.exec(select(DocumentJob).where(DocumentJob.client_quarter_id == victim_quarter_id)).all()
            victim_job_ids = [job.id for job in victim_jobs if job.id is not None]
            victim_docs = (
                session.exec(select(GeneratedDocument).where(GeneratedDocument.document_job_id.in_(victim_job_ids))).all()
                if victim_job_ids
                else []
            )
            victim_emails = session.exec(select(EmailMessage).where(EmailMessage.client_quarter_id == victim_quarter_id)).all()
            survivor_rows = session.exec(select(Counterparty).where(Counterparty.client_id == survivor["id"])).all()

        assert message.startswith("Deleted client: Delete Me.")
        assert victim_client is None
        assert victim_rows == []
        assert victim_templates == []
        assert victim_quarters == []
        assert victim_jobs == []
        assert victim_docs == []
        assert victim_emails == []
        assert len(survivor_rows) == 2
        assert not generated_dir.exists()
        assert not template_dir.exists()
    finally:
        controller.close()


def test_modern_controller_uploaded_docx_and_static_attachments_generate_row_values(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "templated_letter.docx"
    letter_path.write_bytes(
        build_docx_from_paragraphs(
            [
                "Dear {{ party_name }}",
                "Balance {{ balance }} for {{ row.party_name }}",
            ]
        )
    )
    note_path = tmp_path / "templated_note.txt"
    note_path.write_text("Note for {{ party_name }}: {{ balance }}", encoding="utf-8")
    static_pdf_path = tmp_path / "authorisation.pdf"
    static_pdf_path.write_bytes(b"%PDF-1.4 static attachment")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])

        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])
        controller.save_document_templates(quarter_id, f"{letter_path}\n{note_path}\n{static_pdf_path}")
        assert controller.auto_map_variables(quarter_id) == "Mappings valid."
        assert controller.run_generation(quarter_id) == "Generated 6 document job(s); 0 need attention."

        with Session(controller.engine) as session:
            documents = session.exec(select(GeneratedDocument)).all()

        generated_suffixes = {Path(document.file_path).name.split("_", 1)[-1] for document in documents}
        letter_texts = [extract_docx_text(Path(document.file_path)) for document in documents if Path(document.file_path).name.endswith("templated_letter.docx")]
        note_texts = [Path(document.file_path).read_text(encoding="utf-8") for document in documents if Path(document.file_path).name.endswith("templated_note.txt")]
        pdf_paths = [Path(document.file_path) for document in documents if Path(document.file_path).name.endswith("authorisation.pdf")]

        assert generated_suffixes == {"templated_letter.docx", "templated_note.txt", "authorisation.pdf"}
        assert any("Dear Alpha Finance" in text and "Balance 1000" in text for text in letter_texts)
        assert any("Dear Beta Bank" in text and "Balance 2000" in text for text in letter_texts)
        assert "Note for Alpha Finance: 1000" in note_texts
        assert "Note for Beta Bank: 2000" in note_texts
        assert len(pdf_paths) == 2
        assert all(path.read_bytes() == static_pdf_path.read_bytes() for path in pdf_paths)
    finally:
        controller.close()


def test_modern_controller_multi_document_save_validates_pdf_and_text_variables(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}", encoding="utf-8")
    authorisation_pdf = tmp_path / "Authorisation for Direct Balance Confirmation.pdf"
    authorisation_pdf.write_bytes(b"%PDF-1.4\nAuthorization for {{Party Name}}\n%%EOF")

    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])

        message = controller.save_document_templates(quarter_id, f"{letter_path}\n{authorisation_pdf}")
        mapping = controller.auto_map_variables(quarter_id)
        readiness = controller.quarter_workflow_readiness(quarter_id)

        assert "Saved 2 document template" in message
        assert "Party Name" in message
        assert mapping == "Mappings valid."
        assert readiness["document_template_count"] == 2
        assert readiness["document_mapping_errors"] == []
        assert readiness["can_generate_documents"] is True
    finally:
        controller.close()


def test_modern_controller_counterparty_crud_uses_client_schema_and_persists_all_fields(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    letter_path = tmp_path / "letter.txt"
    letter_path.write_text("Letter for {{ party_name }}: {{ balance }}", encoding="utf-8")
    try:
        created = controller.create_client_record("Purple United")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])
        controller.import_workbook(quarter_id, str(workbook_path), client_id=created["id"])

        payload = controller.counterparty_form_payload(created["id"], quarter_id)
        assert "Custom Ref" in payload["columns"]
        assert {"Party Type", "Party Name", "Email To(Address)", "Balance"} <= set(payload["columns"])

        add_message = controller.create_counterparty(
            created["id"],
            quarter_id,
            {
                "Party Type": "Creditor",
                "Party Name": "Delta Holdings",
                "Email To(Address)": "delta@example.com",
                "Balance": "4500",
                "Custom Ref": "D-9",
            },
        )
        rows_after_add = controller.quarter_counterparty_statuses(quarter_id)
        delta_row = next(row for row in rows_after_add if row["party_name"] == "Delta Holdings")
        delta_id = int(delta_row["counterparty_id"])
        edit_payload = controller.counterparty_form_payload(created["id"], quarter_id, delta_id)
        controller.save_document_templates(quarter_id, str(letter_path))
        controller.save_mail_templates(quarter_id, "Confirm {{ party_name }}", "Balance {{ balance }}")
        assert controller.auto_map_variables(quarter_id) == "Mappings valid."
        assert controller.regenerate_documents(quarter_id, {delta_id}).startswith("Regenerated")
        assert controller.queue_and_preview_send_selected(quarter_id, {delta_id}).startswith("Queued 1 selected email(s)")

        update_message = controller.update_counterparty(
            created["id"],
            quarter_id,
            delta_id,
            {
                "Party Type": "Debtor",
                "Party Name": "Delta Revised",
                "Email To(Address)": "delta.revised@example.com",
                "Balance": "5000",
                "Custom Ref": "D-10",
            },
        )
        row_after_update = next(row for row in controller.quarter_counterparty_statuses(quarter_id) if row["counterparty_id"] == delta_id)

        with Session(controller.engine) as session:
            counterparty = session.get(Counterparty, delta_id)
            assert counterparty is not None
            fields = session.exec(select(CounterpartyField).where(CounterpartyField.counterparty_id == delta_id)).all()
            jobs = session.exec(select(DocumentJob).where(DocumentJob.counterparty_id == delta_id)).all()
            emails = session.exec(select(EmailMessage).where(EmailMessage.counterparty_id == delta_id)).all()

        assert add_message.startswith("Added counterparty:")
        assert len(rows_after_add) == 3
        assert edit_payload["values"]["Custom Ref"] == "D-9"
        assert update_message.startswith("Updated counterparty:")
        assert counterparty.party_name == "Delta Revised"
        assert counterparty.email == "delta.revised@example.com"
        assert counterparty.balance == "5000"
        assert counterparty.status == "non_compliant"
        assert {field.field_name for field in fields} >= set(payload["columns"])
        assert next(field for field in fields if field.field_name == "Custom Ref").field_value == "D-10"
        assert row_after_update["document_status"] == "not generated"
        assert row_after_update["mail_status"] == "not generated"
        assert jobs == []
        assert emails == []

        delete_message = controller.delete_counterparty(created["id"], quarter_id, delta_id)
        rows_after_delete = controller.quarter_counterparty_statuses(quarter_id)
        with Session(controller.engine) as session:
            deleted = session.get(Counterparty, delta_id)
            deleted_fields = session.exec(select(CounterpartyField).where(CounterpartyField.counterparty_id == delta_id)).all()

        assert delete_message.startswith("Deleted counterparty:")
        assert deleted is None
        assert deleted_fields == []
        assert [row["party_name"] for row in rows_after_delete] == ["Alpha Finance", "Beta Bank"]
    finally:
        controller.close()


def test_modern_controller_counterparty_crud_is_scoped_to_client_and_quarter(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    workbook_path = _sample_workbook(tmp_path)
    try:
        client_a = controller.create_client_record("Client A")
        controller.create_quarter(client_a["id"], "2026-27", "Q1")
        quarter_a = controller.current_quarter_id_for_client(client_a["id"])
        controller.import_workbook(quarter_a, str(workbook_path), client_id=client_a["id"])

        client_b = controller.create_client_record("Client B")
        controller.create_quarter(client_b["id"], "2026-27", "Q1")
        quarter_b = controller.current_quarter_id_for_client(client_b["id"])
        controller.import_workbook(quarter_b, str(workbook_path), client_id=client_b["id"])

        row_a_id = controller.quarter_counterparty_statuses(quarter_a)[0]["counterparty_id"]
        row_b_id = controller.quarter_counterparty_statuses(quarter_b)[0]["counterparty_id"]

        with pytest.raises(ValueError, match="does not belong to this client"):
            controller.counterparty_form_payload(client_b["id"], quarter_a, int(row_a_id))
        with pytest.raises(ValueError, match="does not belong to the active client quarter"):
            controller.update_counterparty(
                client_b["id"],
                quarter_b,
                int(row_a_id),
                {
                    "Party Type": "Creditor",
                    "Party Name": "Cross Scope",
                    "Email To(Address)": "cross@example.com",
                    "Balance": "1",
                    "Custom Ref": "X-1",
                },
            )
        with pytest.raises(ValueError, match="does not belong to the active client quarter"):
            controller.delete_counterparty(client_a["id"], quarter_a, int(row_b_id))
    finally:
        controller.close()


def test_modern_controller_counterparty_create_requires_excel_schema_import(tmp_path: Path) -> None:
    controller = ModernComplianceController(tmp_path / "modern.db")
    try:
        created = controller.create_client_record("Client A")
        controller.create_quarter(created["id"], "2026-27", "Q1")
        quarter_id = controller.current_quarter_id_for_client(created["id"])

        with pytest.raises(ValueError, match="Import Excel for this quarter"):
            controller.counterparty_form_payload(created["id"], quarter_id)
        with pytest.raises(ValueError, match="Import Excel for this quarter"):
            controller.create_counterparty(
                created["id"],
                quarter_id,
                {
                    "Party Type": "Creditor",
                    "Party Name": "Manual Row",
                    "Email To(Address)": "manual@example.com",
                    "Balance": "100",
                },
            )
    finally:
        controller.close()


def test_document_templates_support_new_excel_columns_without_manual_mapping(tmp_path: Path) -> None:
    engine = init_db(tmp_path / "compliance.db")
    workbook_path = _type_workbook(tmp_path)
    doc_path = tmp_path / "type_confirmation.txt"
    doc_path.write_text("Type {{ Type }} for {{ Party Name }}", encoding="utf-8")

    with Session(engine) as session:
        client, quarter = _client_and_quarter(session)
        ImportService(session).import_excel(client.id, quarter.id, workbook_path)
        template_service = TemplateService(session)
        template_service.save_document_templates(
            quarter.id,
            [doc_path],
            storage_dir=tmp_path / "template_store",
            client_id=client.id,
        )

        # New Excel columns should work directly via {{ Column Name }} placeholders.
        assert template_service.validate_mappings(quarter.id, {"Party Name", "Type", "Email To(Address)", "Balance"}) == []

        workflow = WorkflowService(session, output_root=tmp_path / "generated")
        workflow.enqueue_document_generation(quarter.id)
        generated = workflow.generate_pending_documents(quarter.id)

    assert len(generated) == 1
    assert generated[0].status == "generated"
    text = Path(generated[0].file_path).read_text(encoding="utf-8")
    assert "Type Creditor" in text
    assert "Alpha Finance" in text


def _client_and_quarter(session: Session):
    service = ClientService(session)
    client = service.create_client("Purple United", "listed_org")
    assert client.id is not None
    quarter = service.create_quarter(client.id, "2026-27", "Q1", current=True)
    assert quarter.id is not None
    return client, quarter


def _sample_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Creditors"
    ws.append(["S.No.", "Party Type", "Party Name", "Email To(Address)", "Balance", "Custom Ref"])
    ws.append([1, "Creditor", "Alpha Finance", "alpha@example.com", "1000", "A-1"])
    ws.append([2, "Creditor", "Beta Bank", "beta@example.com", "2000", "B-1"])
    wb.save(workbook_path)
    return workbook_path


def _type_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "type_master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Creditors"
    ws.append(["S.No.", "Type", "Party Name", "Email To(Address)", "Balance"])
    ws.append([1, "Creditor", "Alpha Finance", "alpha@example.com", "1000"])
    wb.save(workbook_path)
    return workbook_path


def _single_row_workbook(tmp_path: Path, filename: str, party_name: str, email: str, balance: str) -> Path:
    workbook_path = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracking"
    ws.append(["S.No.", "Party Type", "Party Name", "Email To(Address)", "Balance"])
    ws.append([1, "Creditor", party_name, email, balance])
    wb.save(workbook_path)
    return workbook_path


def _status_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "status_master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Creditors"
    ws.append(
        [
            "S.No.",
            "Party Type",
            "Party Name",
            "Email To(Address)",
            "Balance",
            "Status",
            "ReadyToSend",
            "BounceReceived",
        ]
    )
    ws.append([1, "Creditor", "Alpha Finance", "alpha@example.com", "1000", "Sent", "", ""])
    ws.append([2, "Creditor", "Beta Bank", "beta@example.com", "2000", "", "Y", ""])
    ws.append([3, "Creditor", "Gamma Capital", "gamma@example.com", "3000", "", "", "Yes"])
    wb.save(workbook_path)
    return workbook_path


def _workbook_with_readme_sheet(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "readme_and_tracking.xlsx"
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme.append(["Instructions"])
    readme.append(["Use the Tracking sheet for creditor/debtor data."])
    readme.append(["Do not import this sheet as workflow rows."])
    tracking = wb.create_sheet("Tracking")
    tracking.append(["S.No.", "Party Type", "Party Name", "Email To(Address)", "Balance"])
    tracking.append([1, "Creditor", "Alpha Finance", "alpha@example.com", "1000"])
    tracking.append([2, "Debtor", "Beta Bank", "beta@example.com", "2000"])
    wb.save(workbook_path)
    return workbook_path
