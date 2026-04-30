from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import ConfirmationRow, DocumentResult, RowState, SendResult, TrackingResult, now_text
from .templates import AUDITOR_REPLY_EMAIL, BALANCE_AS_ON_DATE, SUBJECT_TEMPLATE

MASTER_HEADERS = [
    "S.No.",
    "Party Type",
    "Party Name",
    "Email To(Address)",
    "Balance",
]
EXCLUDED_SHEETS = {"Banks"}

TRACKING_HEADERS = [
    "PartyId",
    "SheetName",
    "S.No.",
    "Party Type",
    "Party Name",
    "Email To(Address)",
    "Balance",
    "Sent Date",
    "Reminder 1",
    "Reminder 2",
    "Reply Recd",
    "AttachmentCreated",
    "GeneratedDocxPaths",
    "GeneratedPdfPaths",
    "AttachmentCreatedAt",
    "ReadyToSend",
    "MainSent",
    "SentDate",
    "GmailMessageId",
    "GmailThreadId",
    "BounceReceived",
    "BounceDate",
    "ReplyReceived",
    "ReplyReceivedDate",
    "LastCheckedAt",
    "AttemptCount",
    "LastAttemptAt",
    "LastSuccessfulStep",
    "NextRetryAt",
    "RetryLocked",
    "BatchId",
    "BatchSequence",
    "VerificationStatus",
    "VerificationErrors",
    "LastLogFile",
    "Status",
    "Error",
    "TemplateVersion",
    "GeneratedAttachmentHash",
]


class ExcelStateStore:
    def __init__(self, master_path: Path, tracking_path: Path | None = None) -> None:
        self.master_path = master_path
        self.work_dir = master_path.parent
        self.tracking_path = tracking_path or self.work_dir / "Tracking.xlsx"
        self.tracking_sheet = "Tracking"

    def validate(self) -> list[str]:
        errors: list[str] = []
        if not self.master_path.exists():
            errors.append(f"Master workbook not found: {self.master_path}")
            return errors
        wb = load_workbook(self.master_path, data_only=True)
        processable_sheets = 0
        for ws in wb.worksheets:
            if ws.title in EXCLUDED_SHEETS or self._is_blank_sheet(ws):
                continue
            processable_sheets += 1
            headers = self._headers(ws)
            missing = sorted(set(MASTER_HEADERS) - set(headers))
            if missing:
                errors.append(f"{ws.title}: missing required master columns: {', '.join(missing)}")
        if processable_sheets == 0:
            errors.append("No non-blank worksheets found to process after excluding Banks")
        return errors

    def ensure_workbooks(self) -> None:
        if self.tracking_path.exists():
            twb = load_workbook(self.tracking_path)
        else:
            twb = Workbook()
            twb.active.title = self.tracking_sheet
        tws = twb[self.tracking_sheet] if self.tracking_sheet in twb.sheetnames else twb.active
        tws.title = self.tracking_sheet
        self._ensure_headers(tws, TRACKING_HEADERS)
        tracking_ids = self._tracking_ids(tws)
        rows = self._master_records()
        for record in rows:
            party_id = record["PartyId"]
            if party_id and party_id not in tracking_ids:
                self._append_tracking_row(tws, record)
        self._save_with_retry(twb, self.tracking_path)

    def load_rows(self) -> list[ConfirmationRow]:
        self.ensure_workbooks()
        twb = load_workbook(self.tracking_path)
        tws = twb[self.tracking_sheet]
        tracking_by_id = self._tracking_records(tws)
        rows: list[ConfirmationRow] = []
        for row_number, record in self._master_records(include_row=True):
            row_id = str(record.get("PartyId") or "").strip()
            if not row_id:
                continue
            state = self._state_from_record(tracking_by_id.get(row_id, {}))
            rows.append(self._row_from_record(row_number, record, state))
        return rows

    def mark_documents_created(self, row_id: str, result: DocumentResult, verification_status: str = "Passed", verification_errors: str = "") -> None:
        updates = {
            "AttachmentCreated": "Y",
            "GeneratedDocxPaths": self._join_paths(result.docx_paths),
            "GeneratedPdfPaths": self._join_paths(result.pdf_paths),
            "AttachmentCreatedAt": result.created_at,
            "ReadyToSend": "Y" if verification_status == "Passed" else "N",
            "VerificationStatus": verification_status,
            "VerificationErrors": verification_errors,
            "Status": "ReadyToSend" if verification_status == "Passed" else "VerificationFailed",
            "Error": verification_errors,
            "TemplateVersion": result.template_version,
            "GeneratedAttachmentHash": result.attachment_hash,
            "LastSuccessfulStep": "document_generation",
        }
        self._update_tracking(row_id, updates)

    def mark_send_success(self, row_id: str, result: SendResult) -> None:
        updates = {
            "MainSent": "Y",
            "SentDate": result.sent_at,
            "Sent Date": result.sent_at,
            "GmailMessageId": result.smtp_message_id,
            "GmailThreadId": result.gmail_thread_id,
            "Status": "Sent",
            "Error": "",
            "LastSuccessfulStep": "gmail_send",
        }
        self._update_tracking(row_id, updates)

    def mark_tracking_update(self, row_id: str, result: TrackingResult) -> None:
        updates = {
            "BounceReceived": "Y" if result.bounce_received else "N",
            "BounceDate": result.bounce_date,
            "ReplyReceived": "Y" if result.reply_received else "N",
            "ReplyReceivedDate": result.reply_received_date,
            "Reply Recd": result.reply_received_date,
            "LastCheckedAt": result.last_checked_at,
            "Status": "ReplyReceived" if result.reply_received else ("Bounced" if result.bounce_received else "Sent"),
            "LastSuccessfulStep": "gmail_tracking",
        }
        self._update_tracking(row_id, updates)

    def mark_batch(self, row_id: str, batch_id: str, sequence: int) -> None:
        self._update_tracking(row_id, {"BatchId": batch_id, "BatchSequence": sequence})

    def mark_log_file(self, row_id: str, log_file: Path) -> None:
        self._update_tracking(row_id, {"LastLogFile": str(log_file)})

    def mark_retryable_error(self, row_id: str, step: str, message: str, next_retry_at: str, locked: bool) -> None:
        current = self._get_tracking_record(row_id)
        attempt_count = int(current.get("AttemptCount") or 0) + 1
        self._update_tracking(
            row_id,
            {
                "AttemptCount": attempt_count,
                "LastAttemptAt": now_text(),
                "NextRetryAt": next_retry_at,
                "RetryLocked": "Y" if locked else "N",
                "Status": "RetryLocked" if locked else "RetryScheduled",
                "Error": f"{step}: {message}",
            },
        )

    def mark_error(self, row_id: str, step: str, message: str) -> None:
        self._update_tracking(
            row_id,
            {
                "LastAttemptAt": now_text(),
                "RetryLocked": "Y",
                "ReadyToSend": "N",
                "Status": "NeedsManualFix",
                "Error": f"{step}: {message}",
            },
        )

    def _row_from_record(self, row_number: int, record: dict[str, object], state: RowState) -> ConfirmationRow:
        name = self._text(record.get("Party Name"))
        first = self._text(record.get("ContactFirstName")) or (name.split(" ")[0] if name else "")
        last = self._text(record.get("ContactLastName")) or (" ".join(name.split(" ")[1:]) if len(name.split(" ")) > 1 else "")
        email = self._text(record.get("Email To(Address)"))
        return ConfirmationRow(
            row_id=self._text(record.get("PartyId")),
            excel_row_number=row_number,
            party=self._text(record.get("Party Type")),
            party_name=name,
            contact_name=name,
            contact_first_name=first,
            contact_last_name=last,
            email=email,
            cc="",
            subject=SUBJECT_TEMPLATE,
            balance=self._text(record.get("Balance")),
            balance_nature="",
            company_name="Purple United Sales Limited",
            address="",
            phone="",
            balance_as_on_date=BALANCE_AS_ON_DATE,
            letter_date=now_text().split(" ")[0],
            auditor_reply_email=AUDITOR_REPLY_EMAIL,
            mail_body_override="",
            extra_attachment_paths=[],
            state=state,
        )

    def _state_from_record(self, record: dict[str, object]) -> RowState:
        return RowState(
            attachment_created=self._text(record.get("AttachmentCreated")) or "N",
            generated_docx_paths=self._split_paths(self._text(record.get("GeneratedDocxPaths"))),
            generated_pdf_paths=self._split_paths(self._text(record.get("GeneratedPdfPaths"))),
            attachment_created_at=self._text(record.get("AttachmentCreatedAt")),
            ready_to_send=self._text(record.get("ReadyToSend")) or "N",
            main_sent=self._text(record.get("MainSent")) or "N",
            sent_date=self._text(record.get("SentDate")),
            gmail_message_id=self._text(record.get("GmailMessageId")),
            gmail_thread_id=self._text(record.get("GmailThreadId")),
            bounce_received=self._text(record.get("BounceReceived")) or "N",
            bounce_date=self._text(record.get("BounceDate")),
            reply_received=self._text(record.get("ReplyReceived")) or "N",
            reply_received_date=self._text(record.get("ReplyReceivedDate")),
            last_checked_at=self._text(record.get("LastCheckedAt")),
            attempt_count=int(record.get("AttemptCount") or 0),
            last_attempt_at=self._text(record.get("LastAttemptAt")),
            last_successful_step=self._text(record.get("LastSuccessfulStep")),
            next_retry_at=self._text(record.get("NextRetryAt")),
            retry_locked=self._text(record.get("RetryLocked")) or "N",
            batch_id=self._text(record.get("BatchId")),
            batch_sequence=self._text(record.get("BatchSequence")),
            verification_status=self._text(record.get("VerificationStatus")),
            verification_errors=self._text(record.get("VerificationErrors")),
            last_log_file=self._text(record.get("LastLogFile")),
            status=self._text(record.get("Status")) or "Pending",
            error=self._text(record.get("Error")),
            template_version=self._text(record.get("TemplateVersion")),
            generated_attachment_hash=self._text(record.get("GeneratedAttachmentHash")),
        )

    def _update_tracking(self, row_id: str, updates: dict[str, object]) -> None:
        wb = load_workbook(self.tracking_path)
        ws = wb[self.tracking_sheet]
        headers = self._headers(ws)
        row_number = self._find_tracking_row(ws, row_id)
        if not row_number:
            raise ValueError(f"Tracking row not found for PartyId {row_id}")
        for key, value in updates.items():
            if key not in headers:
                headers[key] = ws.max_column + 1
                ws.cell(1, headers[key]).value = key
            ws.cell(row_number, headers[key]).value = value
        self._save_with_retry(wb, self.tracking_path)

    def _get_tracking_record(self, row_id: str) -> dict[str, object]:
        wb = load_workbook(self.tracking_path)
        ws = wb[self.tracking_sheet]
        row_number = self._find_tracking_row(ws, row_id)
        if not row_number:
            return {}
        headers = self._headers(ws)
        return {header: ws.cell(row_number, col).value for header, col in headers.items()}

    def _headers(self, ws) -> dict[str, int]:
        return {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}

    def _ensure_headers(self, ws, required: list[str]) -> None:
        headers = self._headers(ws)
        if ws.max_row == 1 and not headers:
            for col, header in enumerate(required, 1):
                ws.cell(1, col).value = header
            return
        for header in required:
            if header not in headers:
                ws.cell(1, ws.max_column + 1).value = header

    def _master_records(self, include_row: bool = False):
        wb = load_workbook(self.master_path, data_only=True)
        records = []
        for ws in wb.worksheets:
            if ws.title in EXCLUDED_SHEETS or self._is_blank_sheet(ws):
                continue
            headers = self._headers(ws)
            for row in range(2, ws.max_row + 1):
                raw_record = {header: ws.cell(row, col).value for header, col in headers.items()}
                if not any(raw_record.get(header) not in (None, "") for header in MASTER_HEADERS):
                    continue
                record = self._normalize_master_record(ws.title, row, raw_record)
                records.append((row, record) if include_row else record)
        return records

    def _tracking_records(self, ws) -> dict[str, dict[str, object]]:
        headers = self._headers(ws)
        records: dict[str, dict[str, object]] = {}
        for row in range(2, ws.max_row + 1):
            record = {header: ws.cell(row, col).value for header, col in headers.items()}
            party_id = self._text(record.get("PartyId"))
            if party_id:
                records[party_id] = record
        return records

    def _tracking_ids(self, ws) -> set[str]:
        return set(self._tracking_records(ws).keys())

    def _append_tracking_row(self, ws, record: dict[str, object]) -> None:
        headers = self._headers(ws)
        row = ws.max_row + 1
        defaults = {
            "PartyId": record.get("PartyId"),
            "SheetName": record.get("SheetName"),
            "S.No.": record.get("S.No."),
            "Party Type": record.get("Party Type"),
            "Party Name": record.get("Party Name"),
            "Email To(Address)": record.get("Email To(Address)"),
            "Balance": record.get("Balance"),
            "AttachmentCreated": "N",
            "ReadyToSend": "N",
            "MainSent": "N",
            "BounceReceived": "N",
            "ReplyReceived": "N",
            "AttemptCount": 0,
            "RetryLocked": "N",
            "Status": "Pending",
        }
        for header, value in defaults.items():
            ws.cell(row, headers[header]).value = value

    def _find_tracking_row(self, ws, row_id: str) -> int | None:
        headers = self._headers(ws)
        col = headers.get("PartyId")
        if not col:
            return None
        for row in range(2, ws.max_row + 1):
            if self._text(ws.cell(row, col).value) == row_id:
                return row
        return None

    def _normalize_master_record(self, sheet_name: str, row_number: int, record: dict[str, object]) -> dict[str, object]:
        serial = self._text(record.get("S.No.")) or str(row_number - 1)
        return {
            "PartyId": f"{self._safe_id(sheet_name)}-{self._safe_id(serial)}",
            "SheetName": sheet_name,
            "S.No.": serial,
            "Party Type": self._text(record.get("Party Type")),
            "Party Name": self._text(record.get("Party Name")),
            "Email To(Address)": self._text(record.get("Email To(Address)")),
            "Balance": self._text(record.get("Balance")),
        }

    def _is_blank_sheet(self, ws) -> bool:
        headers = self._headers(ws)
        if not headers:
            return True
        for row in range(2, ws.max_row + 1):
            for col in headers.values():
                if ws.cell(row, col).value not in (None, ""):
                    return False
        return True

    def _save_with_retry(self, wb, path: Path) -> None:
        try:
            wb.save(path)
        except PermissionError as exc:
            raise PermissionError(f"Workbook is locked/open: {path}") from exc

    def _split_paths(self, value: str) -> list[Path]:
        if not value:
            return []
        parts = [part.strip() for part in re.split(r"[;\n]", value) if part.strip()]
        return [Path(part) for part in parts]

    def _join_paths(self, paths: list[Path]) -> str:
        return "; ".join(str(path) for path in paths)

    def _text(self, value: object) -> str:
        return "" if value is None else str(value).strip()

    def _safe_id(self, value: str) -> str:
        text = re.sub(r"[^A-Za-z0-9]+", "_", value.strip()).strip("_")
        return text or "row"
