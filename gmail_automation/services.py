from __future__ import annotations

import hashlib
import shutil
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlmodel import Session, select

from .dao import AuditDAO, ClientDAO, ImportDAO, TemplateDAO, WorkflowDAO
from .db_models import (
    Client,
    ClientQuarter,
    Counterparty,
    CounterpartyField,
    DocumentJob,
    EmailBatch,
    EmailMessage,
    ExcelImport,
    GeneratedDocument,
    Template,
    VariableMapping,
    WorkflowRun,
    utc_now_text,
)
from .docx_utils import extract_docx_text, replace_docx_text
from .errors import ValidationError
from .liquid_utils import build_nested_context, extract_liquid_variables, render_liquid_template
from .mail_sender import send_with_fallback
from .models import ConfirmationRow, DocumentResult, MailTemplate, RowState, SendConfig
from .retry import RetryPolicy


STANDARD_COLUMN_ALIASES = {
    "party_name": ["Party Name", "Party", "Name", "Lender Name", "Creditor Name", "Debtor Name"],
    "email": ["Email To(Address)", "Email", "Email Address", "Lender Email", "Creditor Email"],
    "party_type": ["Party Type", "Type", "Counterparty Type"],
    "balance": ["Balance", "Amount", "Outstanding Balance"],
}

BUILTIN_CONTEXT_VARIABLES = {
    "party_name",
    "email",
    "party_type",
    "balance",
    "row.party_name",
    "row.email",
    "row.party_type",
    "row.balance",
    "quarter.name",
    "quarter.financial_year",
    "quarter.quarter",
}

STATUS_COLUMNS = (
    "Status",
    "Compliance Status",
    "MainSent",
    "ReadyToSend",
    "ReplyReceived",
    "BounceReceived",
    "AttachmentCreated",
)


@dataclass(frozen=True)
class ImportSummary:
    excel_import_id: int
    rows_imported: int
    detected_columns: list[str]
    sheet_counts: dict[str, int]


@dataclass(frozen=True)
class ComplianceSummary:
    total: int
    non_compliant: int
    partially_compliant: int
    fully_compliant: int
    party_type_counts: dict[str, int]
    generation_counts: dict[str, int]
    email_counts: dict[str, int]


@dataclass(frozen=True)
class GeneratedJobResult:
    job_id: int
    counterparty_id: int
    status: str
    file_path: str = ""
    error: str = ""


class ClientService:
    def __init__(self, session: Session) -> None:
        self.clients = ClientDAO(session)
        self.audit = AuditDAO(session)

    def create_client(self, name: str, client_type: str = "listed_org", **metadata: Any) -> Client:
        client = self.clients.create_client(name=name, client_type=client_type, **metadata)
        self.audit.log("client_created", f"Client created: {name}", entity_type="client", entity_id=str(client.id))
        return client

    def create_quarter(self, client_id: int, financial_year: str, quarter: str, current: bool = True) -> ClientQuarter:
        client_quarter = self.clients.create_quarter(client_id, financial_year, quarter, current)
        self.audit.log(
            "quarter_created",
            f"Quarter created: {financial_year} {quarter}",
            client_quarter_id=client_quarter.id,
            entity_type="client_quarter",
            entity_id=str(client_quarter.id),
        )
        return client_quarter


class ImportService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.imports = ImportDAO(session)
        self.audit = AuditDAO(session)

    def import_excel(
        self,
        client_id: int,
        client_quarter_id: int,
        workbook_path: Path,
        ignored_sheets: set[str] | None = None,
        replace_existing: bool = False,
    ) -> ImportSummary:
        ignored_sheets = ignored_sheets or {"Banks"}
        if not workbook_path.exists():
            raise ValidationError(f"Workbook not found: {workbook_path}")

        workbook_hash = _sha256_file(workbook_path)
        wb = load_workbook(workbook_path, data_only=True)
        detected_columns: set[str] = set()
        sheet_counts: dict[str, int] = {}
        sheet_metadata: dict[str, dict[str, Any]] = {}

        for ws in wb.worksheets:
            if ws.title in ignored_sheets or _is_blank_sheet(ws):
                continue
            headers = _headers(ws)
            if not _has_counterparty_headers(headers):
                continue
            detected_columns.update(headers)
            count = _data_row_count(ws, headers)
            if count:
                sheet_counts[ws.title] = count
                sheet_metadata[ws.title] = {"rows": count, "columns": headers}

        if not sheet_counts:
            raise ValidationError("No importable Excel rows found.")

        if replace_existing:
            self._clear_quarter_imported_workflow_data(client_quarter_id)

        excel_import = self.imports.create_import(
            client_quarter_id=client_quarter_id,
            source_path=workbook_path,
            workbook_hash=workbook_hash,
            sheet_metadata=sheet_metadata,
            detected_columns=sorted(detected_columns),
        )
        if excel_import.id is None:
            raise ValidationError("Excel import was not persisted.")

        rows_imported = 0
        for ws in wb.worksheets:
            if ws.title in ignored_sheets or _is_blank_sheet(ws):
                continue
            headers = _headers(ws)
            if not _has_counterparty_headers(headers):
                continue
            for row_number in range(2, ws.max_row + 1):
                values = {header: _text(ws.cell(row_number, column).value) for header, column in headers.items()}
                if not any(values.values()):
                    continue
                normalized = _with_standard_columns(values)
                if not _has_counterparty_values(normalized):
                    continue
                row_key = f"{_safe_id(ws.title)}-{_safe_id(normalized.get('S.No.') or str(row_number - 1))}"
                self.imports.upsert_counterparty(
                    client_id=client_id,
                    client_quarter_id=client_quarter_id,
                    excel_import_id=excel_import.id,
                    source_sheet=ws.title,
                    source_row_number=row_number,
                    source_row_key=row_key,
                    values=normalized,
                )
                rows_imported += 1

        self.audit.log(
            "excel_imported",
            f"Imported {rows_imported} counterparties from {workbook_path.name}",
            client_quarter_id=client_quarter_id,
            entity_type="excel_import",
            entity_id=str(excel_import.id),
            details={"sheet_counts": sheet_counts, "detected_columns": sorted(detected_columns)},
        )
        return ImportSummary(excel_import.id, rows_imported, sorted(detected_columns), sheet_counts)

    def _clear_quarter_imported_workflow_data(self, client_quarter_id: int) -> None:
        counterparties = list(self.session.exec(select(Counterparty).where(Counterparty.client_quarter_id == client_quarter_id)))
        counterparty_ids = [counterparty.id for counterparty in counterparties if counterparty.id is not None]
        if counterparty_ids:
            document_job_ids = [
                job.id
                for job in self.session.exec(select(DocumentJob).where(DocumentJob.counterparty_id.in_(counterparty_ids)))
                if job.id is not None
            ]
            if document_job_ids:
                self.session.exec(delete(GeneratedDocument).where(GeneratedDocument.document_job_id.in_(document_job_ids)))
            self.session.exec(delete(DocumentJob).where(DocumentJob.counterparty_id.in_(counterparty_ids)))
            self.session.exec(delete(EmailMessage).where(EmailMessage.counterparty_id.in_(counterparty_ids)))
            self.session.exec(delete(CounterpartyField).where(CounterpartyField.counterparty_id.in_(counterparty_ids)))
            self.session.exec(delete(Counterparty).where(Counterparty.id.in_(counterparty_ids)))
        self.session.exec(delete(EmailBatch).where(EmailBatch.client_quarter_id == client_quarter_id))
        self.session.exec(delete(WorkflowRun).where(WorkflowRun.client_quarter_id == client_quarter_id))
        self.session.exec(delete(ExcelImport).where(ExcelImport.client_quarter_id == client_quarter_id))
        self.session.commit()


class TemplateService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.templates = TemplateDAO(session)
        self.audit = AuditDAO(session)

    def save_text_template(
        self,
        client_quarter_id: int,
        template_type: str,
        name: str,
        content_text: str,
        client_id: int | None = None,
    ) -> Template:
        self.templates.deactivate_templates(client_quarter_id, {template_type})
        variables = extract_liquid_variables(content_text)
        checksum = _sha256_text(content_text)
        template = self.templates.create_template(
            template_type=template_type,
            name=name,
            version=checksum[:12],
            checksum=checksum,
            client_id=client_id,
            client_quarter_id=client_quarter_id,
            content_text=content_text,
            variables=variables,
        )
        self.audit.log("template_saved", f"Template saved: {name}", client_quarter_id=client_quarter_id, entity_type="template", entity_id=str(template.id))
        return template

    def save_file_template(
        self,
        client_quarter_id: int,
        template_type: str,
        name: str,
        file_path: Path,
        storage_dir: Path,
        client_id: int | None = None,
    ) -> Template:
        if not file_path.exists():
            raise ValidationError(f"Template file not found: {file_path}")
        storage_dir.mkdir(parents=True, exist_ok=True)
        stored_path = storage_dir / f"{_sha256_file(file_path)[:12]}_{file_path.name}"
        if file_path.resolve() != stored_path.resolve():
            shutil.copy2(file_path, stored_path)

        text = _extract_template_text(stored_path)
        variables = extract_liquid_variables(text)
        checksum = _sha256_file(stored_path)
        template = self.templates.create_template(
            template_type=template_type,
            name=name,
            version=checksum[:12],
            checksum=checksum,
            client_id=client_id,
            client_quarter_id=client_quarter_id,
            file_path=str(stored_path),
            variables=variables,
        )
        self.audit.log("template_saved", f"Template file saved: {name}", client_quarter_id=client_quarter_id, entity_type="template", entity_id=str(template.id))
        return template

    def save_document_templates(
        self,
        client_quarter_id: int,
        file_paths: list[Path],
        storage_dir: Path,
        client_id: int | None = None,
    ) -> list[Template]:
        for file_path in file_paths:
            if not file_path.exists():
                raise ValidationError(f"Template file not found: {file_path}")
        self.templates.deactivate_templates(client_quarter_id, {"document"})
        templates: list[Template] = []
        for file_path in file_paths:
            templates.append(
                self.save_file_template(
                    client_quarter_id=client_quarter_id,
                    template_type="document",
                    name=file_path.stem,
                    file_path=file_path,
                    storage_dir=storage_dir,
                    client_id=client_id,
                )
            )
        return templates

    def required_variables(self, client_quarter_id: int) -> set[str]:
        return self.templates.variables_for_quarter(client_quarter_id)

    def save_mappings(self, client_quarter_id: int, mappings: dict[str, tuple[str, str, str]]) -> None:
        self.templates.save_mappings(client_quarter_id, mappings)
        self.audit.log("mappings_saved", f"Saved {len(mappings)} variable mappings", client_quarter_id=client_quarter_id)

    def validate_mappings(self, client_quarter_id: int, available_columns: set[str]) -> list[str]:
        variables = self.required_variables(client_quarter_id)
        mappings = self.templates.mappings_for_quarter(client_quarter_id)
        errors: list[str] = []
        for variable in sorted(variables):
            if _is_builtin_context_variable(variable):
                continue
            mapping = mappings.get(variable)
            if mapping is None:
                errors.append(f"{variable}: missing mapping")
                continue
            if mapping.source_type == "excel_column" and mapping.source_key not in available_columns:
                errors.append(f"{variable}: Excel column not found: {mapping.source_key}")
            elif mapping.source_type == "constant" and not mapping.constant_value:
                errors.append(f"{variable}: constant value is empty")
        return errors

    def build_context(self, client_quarter_id: int, counterparty: Counterparty, row_fields: dict[str, str]) -> dict[str, Any]:
        mappings = self.templates.mappings_for_quarter(client_quarter_id)
        quarter = self.session.get(ClientQuarter, client_quarter_id)
        quarter_name = f"{quarter.financial_year} {quarter.quarter}" if quarter is not None else ""
        flat_values: dict[str, str] = {
            "row.party_name": counterparty.party_name,
            "row.email": counterparty.email,
            "row.party_type": counterparty.party_type,
            "row.balance": counterparty.balance,
            "party_name": counterparty.party_name,
            "email": counterparty.email,
            "party_type": counterparty.party_type,
            "balance": counterparty.balance,
            "quarter.name": quarter_name,
            "quarter.financial_year": quarter.financial_year if quarter is not None else "",
            "quarter.quarter": quarter.quarter if quarter is not None else "",
        }
        for variable, mapping in mappings.items():
            flat_values[variable] = _resolve_mapping(mapping, row_fields, counterparty)
        return build_nested_context(flat_values)


class DashboardService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.imports = ImportDAO(session)
        self.workflow = WorkflowDAO(session)

    def compliance_summary(self, client_quarter_id: int) -> ComplianceSummary:
        ComplianceService(self.session).reconcile_quarter(client_quarter_id)
        counterparties = self.imports.list_counterparties(client_quarter_id)
        counts = Counter()
        party_type_counts = Counter((counterparty.party_type or "Unspecified") for counterparty in counterparties)
        for counterparty in counterparties:
            counts["fully_compliant" if counterparty.status == "compliant" else "non_compliant"] += 1

        return ComplianceSummary(
            total=len(counterparties),
            non_compliant=counts["non_compliant"],
            partially_compliant=0,
            fully_compliant=counts["fully_compliant"],
            party_type_counts=dict(party_type_counts),
            generation_counts=dict(self.workflow.generated_document_counts(client_quarter_id)),
            email_counts=dict(self.workflow.email_status_counts(client_quarter_id)),
        )


class ComplianceService:
    def __init__(self, session: Session) -> None:
        self.session = session

    def reset_quarter_status(self, client_quarter_id: int, reset_documents: bool = False, reset_mail: bool = False) -> int:
        counterparties = list(self.session.exec(select(Counterparty).where(Counterparty.client_quarter_id == client_quarter_id)))
        counterparty_ids = [counterparty.id for counterparty in counterparties if counterparty.id is not None]
        if not counterparty_ids:
            return 0

        if reset_documents:
            document_job_ids = [
                job.id
                for job in self.session.exec(select(DocumentJob).where(DocumentJob.client_quarter_id == client_quarter_id))
                if job.id is not None
            ]
            if document_job_ids:
                self.session.exec(delete(GeneratedDocument).where(GeneratedDocument.document_job_id.in_(document_job_ids)))
            self.session.exec(delete(GeneratedDocument).where(GeneratedDocument.counterparty_id.in_(counterparty_ids)))
            self.session.exec(delete(DocumentJob).where(DocumentJob.client_quarter_id == client_quarter_id))

        if reset_documents or reset_mail:
            self.session.exec(delete(EmailMessage).where(EmailMessage.client_quarter_id == client_quarter_id))
            self.session.exec(delete(EmailBatch).where(EmailBatch.client_quarter_id == client_quarter_id))

        for counterparty in counterparties:
            counterparty.status = "non_compliant"
            counterparty.updated_at = utc_now_text()
            self.session.add(counterparty)
        self.session.commit()
        return len(counterparties)

    def reconcile_all(self) -> int:
        changed = 0
        quarter_ids = [
            quarter.id
            for quarter in self.session.exec(select(ClientQuarter))
            if quarter.id is not None
        ]
        for quarter_id in quarter_ids:
            changed += self.reconcile_quarter(quarter_id)
        return changed

    def reconcile_quarter(self, client_quarter_id: int) -> int:
        counterparties = list(self.session.exec(select(Counterparty).where(Counterparty.client_quarter_id == client_quarter_id)))
        if not counterparties:
            return 0
        counterparty_ids = [counterparty.id for counterparty in counterparties if counterparty.id is not None]
        if not counterparty_ids:
            return 0

        quarter_job_ids = [
            job.id
            for job in self.session.exec(
                select(DocumentJob).where(
                    DocumentJob.client_quarter_id == client_quarter_id,
                    DocumentJob.counterparty_id.in_(counterparty_ids),
                )
            )
            if job.id is not None
        ]
        generated_ids = set()
        if quarter_job_ids:
            generated_ids = {
                counterparty_id
                for counterparty_id in self.session.exec(
                    select(GeneratedDocument.counterparty_id).where(GeneratedDocument.document_job_id.in_(quarter_job_ids))
                )
            }
        sent_ids = {
            message.counterparty_id
            for message in self.session.exec(
                select(EmailMessage).where(
                    EmailMessage.client_quarter_id == client_quarter_id,
                    EmailMessage.counterparty_id.in_(counterparty_ids),
                    EmailMessage.status == "sent",
                )
            )
            if not (message.smtp_message_id or "").startswith("preview-")
        }

        changed = 0
        for counterparty in counterparties:
            if counterparty.id is None:
                continue
            expected_status = "compliant" if counterparty.id in generated_ids and counterparty.id in sent_ids else "non_compliant"
            if counterparty.status != expected_status:
                counterparty.status = expected_status
                counterparty.updated_at = utc_now_text()
                self.session.add(counterparty)
                changed += 1
        if changed:
            self.session.commit()
        return changed


class WorkflowService:
    def __init__(self, session: Session, config: SendConfig | None = None, output_root: Path | None = None) -> None:
        self.session = session
        self.config = config or SendConfig()
        self.output_root = output_root or Path.cwd() / "generated"
        self.imports = ImportDAO(session)
        self.templates = TemplateDAO(session)
        self.template_service = TemplateService(session)
        self.workflow = WorkflowDAO(session)
        self.audit = AuditDAO(session)
        self.retry_policy = RetryPolicy()

    def enqueue_document_generation(self, client_quarter_id: int) -> list[DocumentJob]:
        counterparties = self.imports.list_counterparties(client_quarter_id)
        document_templates = [
            template
            for template in self.templates.list_templates(client_quarter_id)
            if template.template_type == "document"
        ]
        run = self.workflow.create_run(client_quarter_id, "document_generation", total_count=len(counterparties) * len(document_templates))
        jobs = self.workflow.create_document_jobs(client_quarter_id, workflow_run_id=run.id)
        self.audit.log("generation_enqueued", f"Queued {len(jobs)} document job(s)", client_quarter_id=client_quarter_id, entity_type="workflow_run", entity_id=str(run.id))
        return jobs

    def generate_pending_documents(self, client_quarter_id: int) -> list[GeneratedJobResult]:
        results: list[GeneratedJobResult] = []
        for job in self.workflow.pending_document_jobs(client_quarter_id):
            if job.id is None:
                continue
            try:
                result = self._generate_document_job(job)
                results.append(result)
            except Exception as exc:
                decision = self.retry_policy.next_attempt("document_generation", job.attempts + 1, True)
                job.attempts += 1
                job.next_retry_at = decision.next_retry_at
                job.retry_locked = decision.locked
                job.status = "retry_scheduled" if decision.should_retry else "failed"
                job.error = str(exc)
                job.updated_at = utc_now_text()
                self.session.add(job)
                self.session.commit()
                results.append(GeneratedJobResult(job.id, job.counterparty_id, job.status, error=str(exc)))
        ComplianceService(self.session).reconcile_quarter(client_quarter_id)
        return results

    def regenerate_documents(self, client_quarter_id: int, counterparty_ids: set[int] | None = None) -> list[GeneratedJobResult]:
        counterparties = self._selected_counterparties(client_quarter_id, counterparty_ids)
        document_templates = [
            template
            for template in self.templates.list_templates(client_quarter_id)
            if template.template_type == "document"
        ]
        run = self.workflow.create_run(client_quarter_id, "document_regeneration", total_count=len(counterparties) * len(document_templates))
        jobs: list[DocumentJob] = []
        selected_counterparty_ids = [counterparty.id for counterparty in counterparties if counterparty.id is not None]
        if selected_counterparty_ids:
            old_job_ids = [
                job.id
                for job in self.session.exec(
                    select(DocumentJob).where(
                        DocumentJob.client_quarter_id == client_quarter_id,
                        DocumentJob.counterparty_id.in_(selected_counterparty_ids),
                    )
                )
                if job.id is not None
            ]
            if old_job_ids:
                self.session.exec(delete(GeneratedDocument).where(GeneratedDocument.document_job_id.in_(old_job_ids)))
            self.session.exec(delete(GeneratedDocument).where(GeneratedDocument.counterparty_id.in_(selected_counterparty_ids)))
            self.session.exec(
                delete(DocumentJob).where(
                    DocumentJob.client_quarter_id == client_quarter_id,
                    DocumentJob.counterparty_id.in_(selected_counterparty_ids),
                )
            )
            self.session.commit()
        for counterparty in counterparties:
            assert counterparty.id is not None
            for template in document_templates:
                assert template.id is not None
                job = DocumentJob(
                    workflow_run_id=run.id,
                    client_quarter_id=client_quarter_id,
                    counterparty_id=counterparty.id,
                    template_id=template.id,
                )
                self.session.add(job)
                jobs.append(job)
        self.session.commit()

        results: list[GeneratedJobResult] = []
        for job in jobs:
            if job.id is None:
                continue
            try:
                results.append(self._generate_document_job(job))
            except Exception as exc:
                decision = self.retry_policy.next_attempt("document_generation", job.attempts + 1, True)
                job.attempts += 1
                job.next_retry_at = decision.next_retry_at
                job.retry_locked = decision.locked
                job.status = "retry_scheduled" if decision.should_retry else "failed"
                job.error = str(exc)
                job.updated_at = utc_now_text()
                self.session.add(job)
                self.session.commit()
                results.append(GeneratedJobResult(job.id, job.counterparty_id, job.status, error=str(exc)))
        ComplianceService(self.session).reconcile_quarter(client_quarter_id)
        self.audit.log("documents_regenerated", f"Regenerated {len(results)} document job(s)", client_quarter_id=client_quarter_id)
        return results

    def queue_email_messages(
        self,
        client_quarter_id: int,
        subject_template: str,
        body_template: str,
        counterparty_ids: set[int] | None = None,
    ) -> list[EmailMessage]:
        counterparties = self._selected_counterparties(client_quarter_id, counterparty_ids)
        batch = self.workflow.create_email_batch(client_quarter_id, batch_key=f"batch_{utc_now_text().replace(' ', '_').replace(':', '')}", batch_size=len(counterparties))
        messages: list[EmailMessage] = []
        for counterparty in counterparties:
            assert counterparty.id is not None
            fields = self.imports.fields_for_counterparty(counterparty.id)
            context = self.template_service.build_context(client_quarter_id, counterparty, fields)
            subject = render_liquid_template(subject_template, context).text
            body = render_liquid_template(body_template, context).text
            messages.append(
                self.workflow.upsert_email_message(
                    client_quarter_id=client_quarter_id,
                    counterparty_id=counterparty.id,
                    to_email=counterparty.email,
                    subject=subject,
                    body=body,
                    email_batch_id=batch.id,
                )
            )
        self.audit.log("emails_queued", f"Queued {len(messages)} email(s)", client_quarter_id=client_quarter_id, entity_type="email_batch", entity_id=str(batch.id))
        return messages

    def mark_preview_batch_sent(self, client_quarter_id: int, counterparty_ids: set[int] | None = None) -> int:
        return self.send_queued_emails(client_quarter_id, preview=True, counterparty_ids=counterparty_ids)

    def send_queued_emails(self, client_quarter_id: int, preview: bool | None = None, counterparty_ids: set[int] | None = None) -> int:
        preview = self.config.send_mode != "send" if preview is None else preview
        statement = select(EmailMessage).where(EmailMessage.client_quarter_id == client_quarter_id, EmailMessage.status == "queued")
        if counterparty_ids is not None:
            statement = statement.where(EmailMessage.counterparty_id.in_(counterparty_ids))
        messages = self.session.exec(statement)
        sent_count = 0
        for message in messages:
            if message.id is None:
                continue
            if sent_count >= self.config.daily_send_limit:
                break
            try:
                if preview:
                    self.workflow.mark_email_preview_sent(message.id, smtp_message_id=f"preview-{message.id}")
                else:
                    result = self._send_email_message(message)
                    self.workflow.mark_email_sent(message.id, result.smtp_message_id, result.gmail_thread_id)
                sent_count += 1
            except Exception as exc:
                decision = self.retry_policy.next_attempt("mail_send", message.attempts + 1, True)
                message.attempts += 1
                message.next_retry_at = decision.next_retry_at
                message.retry_locked = decision.locked
                message.status = "retry_scheduled" if decision.should_retry else "failed"
                message.error = str(exc)
                message.updated_at = utc_now_text()
                self.session.add(message)
                self.session.commit()
        ComplianceService(self.session).reconcile_quarter(client_quarter_id)
        event_type = "emails_preview_sent" if preview else "emails_sent"
        self.audit.log(event_type, f"Processed {sent_count} queued email(s)", client_quarter_id=client_quarter_id)
        return sent_count

    def _selected_counterparties(self, client_quarter_id: int, counterparty_ids: set[int] | None = None) -> list[Counterparty]:
        counterparties = self.imports.list_counterparties(client_quarter_id)
        if counterparty_ids is None:
            return counterparties
        return [counterparty for counterparty in counterparties if counterparty.id in counterparty_ids]

    def _send_email_message(self, message: EmailMessage):
        counterparty = self.session.get(Counterparty, message.counterparty_id)
        if counterparty is None:
            raise ValidationError(f"Counterparty not found: {message.counterparty_id}")
        assert counterparty.id is not None
        quarter_job_ids = [
            job.id
            for job in self.session.exec(
                select(DocumentJob).where(
                    DocumentJob.client_quarter_id == message.client_quarter_id,
                    DocumentJob.counterparty_id == counterparty.id,
                )
            )
            if job.id is not None
        ]
        documents = (
            list(self.session.exec(select(GeneratedDocument).where(GeneratedDocument.document_job_id.in_(quarter_job_ids))))
            if quarter_job_ids
            else []
        )
        if not documents:
            raise ValidationError(f"No generated documents found for {counterparty.party_name}")
        row = ConfirmationRow(
            row_id=str(counterparty.id),
            excel_row_number=counterparty.source_row_number,
            party=counterparty.party_type,
            party_name=counterparty.party_name,
            contact_name=counterparty.party_name,
            contact_first_name=counterparty.party_name.split(" ")[0] if counterparty.party_name else "",
            contact_last_name=" ".join(counterparty.party_name.split(" ")[1:]) if len(counterparty.party_name.split(" ")) > 1 else "",
            email=message.to_email,
            cc=message.cc_email,
            subject=message.subject,
            balance=counterparty.balance,
            balance_nature="",
            company_name="",
            address="",
            phone="",
            balance_as_on_date="",
            letter_date=utc_now_text().split(" ")[0],
            auditor_reply_email="",
            mail_body_override=message.body,
            extra_attachment_paths=[],
            state=RowState(ready_to_send="Y", verification_status="Passed"),
        )
        document_paths = [Path(document.file_path) for document in documents]
        document_result = DocumentResult(
            docx_paths=[path for path in document_paths if path.suffix.lower() == ".docx"],
            pdf_paths=[path for path in document_paths if path.suffix.lower() == ".pdf"],
            extra_attachment_paths=[path for path in document_paths if path.suffix.lower() not in {".docx", ".pdf"}],
            created_at=utc_now_text(),
            template_version="db",
            attachment_hash=_sha256_text("|".join(sorted(str(path) for path in document_paths))),
        )
        mail_template = MailTemplate(subject_template=message.subject, body_template_text=message.body, required_fields=set())
        return send_with_fallback(self.config, row, mail_template, document_result)

    def _generate_document_job(self, job: DocumentJob) -> GeneratedJobResult:
        counterparty = self.session.get(Counterparty, job.counterparty_id)
        if counterparty is None:
            raise ValidationError(f"Counterparty not found: {job.counterparty_id}")
        template = self.session.get(Template, job.template_id) if job.template_id is not None else None
        if template is None:
            raise ValidationError(f"Template not found for job: {job.id}")
        assert job.id is not None
        assert counterparty.id is not None

        row_fields = self.imports.fields_for_counterparty(counterparty.id)
        context = self.template_service.build_context(job.client_quarter_id, counterparty, row_fields)
        output_dir = self.output_root / str(job.client_quarter_id) / str(counterparty.id)
        output_dir.mkdir(parents=True, exist_ok=True)

        if template.file_path and Path(template.file_path).suffix.lower() == ".docx":
            output_path = output_dir / Path(template.file_path).name
            replacements = _liquid_replacements(template.file_path, context)
            rendered = replace_docx_text(Path(template.file_path).read_bytes(), replacements)
            output_path.write_bytes(rendered)
        elif template.file_path and Path(template.file_path).suffix.lower() in {".txt", ".html", ".htm", ".md"}:
            source = Path(template.file_path)
            text = source.read_text(encoding="utf-8", errors="ignore")
            output_path = output_dir / source.name
            output_path.write_text(render_liquid_template(text, context).text, encoding="utf-8")
        elif template.file_path:
            source = Path(template.file_path)
            output_path = output_dir / source.name
            output_path.write_bytes(source.read_bytes())
        else:
            output_path = output_dir / f"{_safe_id(template.name)}.txt"
            output_path.write_text(render_liquid_template(template.content_text, context).text, encoding="utf-8")

        generated = GeneratedDocument(
            document_job_id=job.id,
            counterparty_id=counterparty.id,
            template_id=template.id,
            file_type=output_path.suffix.lstrip(".") or "txt",
            file_path=str(output_path),
            checksum=_sha256_file(output_path),
            template_version=template.version,
        )
        self.session.add(generated)
        self.workflow.mark_document_job(job.id, "generated")
        self.audit.log("document_generated", f"Generated document: {output_path.name}", client_quarter_id=job.client_quarter_id, counterparty_id=counterparty.id, entity_type="document_job", entity_id=str(job.id))
        return GeneratedJobResult(job.id, counterparty.id, "generated", str(output_path))


def _liquid_replacements(file_path: str, context: dict[str, Any]) -> dict[str, str]:
    variables = extract_liquid_variables(extract_docx_text(Path(file_path)))
    replacements: dict[str, str] = {}
    for variable in variables:
        rendered = render_liquid_template("{{ " + variable + " }}", context).text
        replacements["{{ " + variable + " }}"] = rendered
        replacements["{{" + variable + "}}"] = rendered
    return replacements


def _extract_template_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return extract_docx_text(path)
    if suffix in {".txt", ".html", ".htm", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".pdf":
        return _extract_pdf_template_text(path)
    return ""


def _extract_pdf_template_text(path: Path) -> str:
    # This validates placeholders in simple/uncompressed PDFs without adding a heavy PDF dependency.
    data = path.read_bytes()
    chunks = [
        data.decode("utf-8", errors="ignore"),
        data.decode("latin-1", errors="ignore"),
    ]
    return "\n".join(chunk for chunk in chunks if "{{" in chunk and "}}" in chunk)


def _is_builtin_context_variable(variable: str) -> bool:
    return variable in BUILTIN_CONTEXT_VARIABLES


def _resolve_mapping(mapping: VariableMapping, row_fields: dict[str, str], counterparty: Counterparty) -> str:
    if mapping.source_type == "constant":
        return mapping.constant_value
    if mapping.source_type == "excel_column":
        return row_fields.get(mapping.source_key, "")
    if mapping.source_type == "counterparty_field":
        return str(getattr(counterparty, mapping.source_key, ""))
    return ""


def _headers(ws) -> dict[str, int]:
    return {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}


def _has_counterparty_headers(headers: dict[str, int]) -> bool:
    header_names = set(headers)
    for aliases in STANDARD_COLUMN_ALIASES.values():
        if header_names.intersection(aliases):
            return True
    return False


def _has_counterparty_values(values: dict[str, str]) -> bool:
    return any(values.get(column) for column in ("Party Name", "Email To(Address)", "Balance"))


def _data_row_count(ws, headers: dict[str, int]) -> int:
    count = 0
    for row in range(2, ws.max_row + 1):
        values = {header: _text(ws.cell(row, column).value) for header, column in headers.items()}
        if any(values.values()) and _has_counterparty_values(_with_standard_columns(values)):
            count += 1
    return count


def _is_blank_sheet(ws) -> bool:
    headers = _headers(ws)
    return not headers or _data_row_count(ws, headers) == 0


def _with_standard_columns(values: dict[str, str]) -> dict[str, str]:
    normalized = dict(values)
    for standard_key, aliases in STANDARD_COLUMN_ALIASES.items():
        canonical = {
            "party_name": "Party Name",
            "email": "Email To(Address)",
            "party_type": "Party Type",
            "balance": "Balance",
        }[standard_key]
        if normalized.get(canonical):
            continue
        for alias in aliases:
            if values.get(alias):
                normalized[canonical] = values[alias]
                break
    imported_status = _imported_compliance_status(values)
    if imported_status:
        normalized["Imported Compliance Status"] = imported_status
    return normalized


def _imported_compliance_status(values: dict[str, str]) -> str:
    status_value = _first_value(values, ("Compliance Status", "Status"))
    normalized = _status_from_text(status_value)
    if normalized:
        return normalized

    if _truthy(values.get("BounceReceived", "")):
        return "non_compliant"
    if _truthy(values.get("ReplyReceived", "")) or _truthy(values.get("MainSent", "")):
        return "compliant"
    return ""


def _first_value(values: dict[str, str], columns: tuple[str, ...]) -> str:
    for column in columns:
        if values.get(column):
            return values[column]
    return ""


def _status_from_text(value: str) -> str:
    text = value.strip().lower().replace("-", "_").replace(" ", "_")
    if not text:
        return ""
    if text in {"fully_compliant", "full", "complete", "completed", "compliant", "sent", "mail_sent", "main_sent", "reply_received", "yes", "y", "true"}:
        return "compliant"
    if text in {
        "partially_compliant",
        "partial",
        "in_progress",
        "generated",
        "attachment_created",
        "ready",
        "ready_to_send",
        "queued",
        "non_compliant",
        "not_compliant",
        "pending",
        "not_started",
        "failed",
        "bounce",
        "bounced",
        "bounce_received",
        "no",
        "n",
        "false",
    }:
        return "non_compliant"
    return ""


def _truthy(value: str) -> bool:
    return value.strip().lower() in {"1", "y", "yes", "true", "sent", "done", "complete", "completed", "received", "created", "ready"}


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _safe_id(value: str) -> str:
    text = "".join(char if char.isalnum() else "_" for char in value.strip()).strip("_")
    return text or "row"


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()
